"""
Generador de plantilla Excel con URLs de WhatsApp para contacto con negocios.
"""

import json
from pathlib import Path
from typing import Mapping
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from autobots.leads.models import Lead, Niche
from autobots.outreach.whatsapp_links import generate_wa_me_link
from autobots.utils.phone import normalize_paraguay_phone_digits


PROJECT_ROOT = Path(__file__).resolve().parents[3]


def _lead_value(lead: Lead | Mapping[str, object], field: str) -> str:
    """Read a string field from a Lead object or a dictionary."""
    if isinstance(lead, Lead):
        return str(getattr(lead, field, "") or "")
    return str(lead.get(field, "") or "")


def generate_outreach_message(lead: Lead | Mapping[str, object], niche: Niche) -> str:
    """Generate a short Spanish outreach message for manual WhatsApp contact."""
    name = _lead_value(lead, "name").strip()
    business_name = name or "tu negocio"

    templates: dict[Niche, str] = {
        "real_estate": (
            "Buenas! Soy Nicolás. Vi {business_name} y estoy ayudando a inmobiliarias "
            "a responder más rápido consultas por WhatsApp, filtrar interesados por zona, "
            "presupuesto y tipo de propiedad, y avisar cuando un lead está listo para seguimiento. "
            "Te puedo mostrar una demo corta?"
        ),
        "retail": (
            "Buenas! Soy Nicolás. Vi {business_name} y estoy ayudando a tiendas a responder "
            "consultas repetitivas por WhatsApp sobre precio, talle, color, stock y delivery. "
            "Te puedo mostrar una idea rápida para vender con menos desorden?"
        ),
        "clinics": (
            "Buenas! Soy Nicolás. Vi {business_name} y estoy ayudando a clínicas y consultorios "
            "a ordenar consultas de turnos por WhatsApp, pedir datos básicos y avisar cuando "
            "alguien quiere agendar. Te puedo mostrar cómo funcionaría?"
        ),
        "beauty": (
            "Buenas! Soy Nicolás. Vi {business_name} y estoy ayudando a salones y barberías "
            "a responder consultas de precios, servicios y turnos por WhatsApp, sin perder "
            "clientes por demora. Te puedo mostrar una demo corta?"
        ),
    }

    return templates[niche].format(business_name=business_name)


def limpiar_telefono(telefono):
    """Limpia el número de teléfono para formato WhatsApp."""
    return normalize_paraguay_phone_digits(telefono)


def generar_url_whatsapp(telefono, nombre_negocio):
    """Genera URL de WhatsApp con mensaje personalizado."""
    mensaje = f"Buenas! Soy Nicolás. Hablo con el responsable de {nombre_negocio}? Vi su local en Google Maps y vi que hay algo que puede hacer que mas clientes les encuentren. Puedo comentarles en unos 30 segundos?"
    
    return generate_wa_me_link(telefono, mensaje)


def generar_plantilla_excel(leads_file=None, output_file=None):
    """Genera archivo Excel con datos de contacto y URLs de WhatsApp."""
    
    # Cargar datos de leads
    leads_file = Path(leads_file) if leads_file else PROJECT_ROOT / "data" / "processed" / "leads.json"
    output_file = Path(output_file) if output_file else PROJECT_ROOT / "data" / "processed" / "plantilla_whatsapp_contacto.xlsx"
    
    with open(leads_file, 'r', encoding='utf-8') as f:
        leads = json.load(f)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos WhatsApp"
    
    # Encabezados
    headers = ['ID', 'Negocio', 'Teléfono Original', 'Teléfono WhatsApp', 'URL WhatsApp', 'Ciudad', 'Rating', 'Reviews', 'Estado']
    ws.append(headers)
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Agregar datos
    row_num = 2
    leads_con_telefono = 0
    leads_sin_telefono = 0
    
    for lead in leads:
        telefono_original = lead.get('telefono', '')
        telefono_limpio = limpiar_telefono(telefono_original)
        
        if telefono_limpio:
            leads_con_telefono += 1
            nombre = lead.get('nombre', '')
            url_whatsapp = generar_url_whatsapp(telefono_limpio, nombre)
            
            # Agregar fila
            ws.cell(row=row_num, column=1, value=lead.get('id'))
            ws.cell(row=row_num, column=2, value=nombre)
            ws.cell(row=row_num, column=3, value=telefono_original)
            ws.cell(row=row_num, column=4, value=telefono_limpio)
            
            # Crear hipervínculo para la URL de WhatsApp
            cell_url = ws.cell(row=row_num, column=5, value=url_whatsapp)
            cell_url.hyperlink = url_whatsapp
            cell_url.font = Font(color="0563C1", underline="single")
            
            ws.cell(row=row_num, column=6, value=lead.get('ciudad'))
            ws.cell(row=row_num, column=7, value=lead.get('rating'))
            ws.cell(row=row_num, column=8, value=lead.get('reviews'))
            ws.cell(row=row_num, column=9, value=lead.get('estado', 'pendiente'))
            
            row_num += 1
        else:
            leads_sin_telefono += 1
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 8,   # ID
        'B': 40,  # Negocio
        'C': 18,  # Teléfono Original
        'D': 18,  # Teléfono WhatsApp
        'E': 80,  # URL WhatsApp
        'F': 15,  # Ciudad
        'G': 10,  # Rating
        'H': 10,  # Reviews
        'I': 12   # Estado
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Guardar archivo
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    print(f"✅ Plantilla generada exitosamente: {output_file}")
    print(f"📊 Total de leads con teléfono: {leads_con_telefono}")
    print(f"⚠️  Total de leads sin teléfono: {leads_sin_telefono}")
    print(f"📱 Total de URLs de WhatsApp generadas: {leads_con_telefono}")
    

if __name__ == "__main__":
    generar_plantilla_excel()
